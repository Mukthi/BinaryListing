import os
from openpyxl import Workbook


extensions = {'.txt':0, '.jar':0, '.py':0}

txt = []
jar = []
py = []

for root, dirs, files in os.walk(".", topdown=False):
	for name in files:
		if name.endswith('.txt'):
			extensions['.txt'] = extensions.get('.txt')+1
			txt.append(os.path.join(root, name))
		if name.endswith('.jar'):
			extensions['.jar'] = extensions.get('.jar')+1
			jar.append(os.path.join(root, name))
		if name.endswith('.py'):
			extensions['.py'] = extensions.get('.py')+1
			py.append(os.path.join(root, name))

print extensions


wb = Workbook()
ws = wb.create_sheet('txt')
row, column = 1, 1
for files in txt:
	ws.cell(row=row, column=column).value = files
	row+=1

ws = wb.create_sheet('jar')
row, column = 1, 1
for files in jar:
	ws.cell(row=row, column=column).value = files
	row+=1

ws = wb.create_sheet('py')
row, column = 1, 1
for files in py:
	ws.cell(row=row, column=column).value = files
	row+=1

wb.save("binary_listing.xlsx")
